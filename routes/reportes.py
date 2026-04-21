import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from routes.decorators import permiso_required
from models import db, Compra, ItemCompra, VentaDiaria, ItemVenta, Asistencia, Empleado
from datetime import datetime, date, timedelta
import pytz

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')
PERU_TZ = pytz.timezone('America/Lima')

def now_peru():
    return datetime.now(PERU_TZ).replace(tzinfo=None)

# ──────────────────────────────────────
#  DASHBOARD DE REPORTES
# ──────────────────────────────────────
@reportes_bp.route('/')
@login_required
def index():
    hoy = date.today()
    return render_template('reportes/index.html', hoy=hoy)


# ──────────────────────────────────────
#  REPORTE ECONÓMICO — similar al Excel
# ──────────────────────────────────────
@reportes_bp.route('/economico')
@login_required
@permiso_required('compras')
def economico():
    hoy = date.today()
    mes_str  = request.args.get('mes',  hoy.strftime('%Y-%m'))
    try:
        anio, mes = int(mes_str.split('-')[0]), int(mes_str.split('-')[1])
    except:
        anio, mes = hoy.year, hoy.month

    inicio = date(anio, mes, 1)
    fin    = date(anio, mes+1, 1) - timedelta(days=1) if mes < 12 else date(anio, 12, 31)

    compras = Compra.query.filter(
        Compra.fecha >= inicio, Compra.fecha <= fin,
        Compra.estado != 'anulado'
    ).order_by(Compra.fecha).all()

    # Construir filas de detalle (estilo registro de compras del Excel)
    filas = []
    item_num = 1
    for c in compras:
        items = c.items
        if not items:
            filas.append({
                'num': item_num, 'fecha': c.fecha,
                'detalle': c.proveedor_nombre or '—',
                'unidad': '', 'cantidad': None, 'precio_unit': None,
                'subtotal': c.total, 'total': c.total,
                'responsable': c.usuario.nombre_completo if c.usuario else '—',
                'comprobante': f"{c.tipo_comprobante or ''} {c.serie_comprobante or ''}-{c.numero_comprobante or ''}".strip(' -'),
                'razon_social': c.proveedor_nombre or ''
            })
            item_num += 1
        else:
            primer = True
            total_compra = c.total
            for it in items:
                filas.append({
                    'num': item_num if primer else '',
                    'fecha': c.fecha if primer else None,
                    'detalle': it.descripcion,
                    'unidad': it.unidad or '',
                    'cantidad': it.cantidad,
                    'precio_unit': it.precio_unitario,
                    'subtotal': it.subtotal,
                    'total': total_compra if primer else None,
                    'responsable': c.usuario.nombre_completo if c.usuario and primer else '',
                    'comprobante': f"{c.tipo_comprobante or ''} {c.serie_comprobante or ''}-{c.numero_comprobante or ''}".strip(' -') if primer else '',
                    'razon_social': c.proveedor_nombre or '' if primer else ''
                })
                primer = False
                item_num += 1

    total_periodo = sum(c.total for c in compras)

    # Ventas del período (para el cuadro de ingresos)
    ventas = VentaDiaria.query.filter(
        VentaDiaria.fecha >= inicio, VentaDiaria.fecha <= fin
    ).all()
    total_ingresos = sum(v.total for v in ventas)

    # Meses disponibles
    meses = []
    d = date(hoy.year, hoy.month, 1)
    for _ in range(12):
        meses.append(d)
        d = date(d.year if d.month > 1 else d.year-1, d.month-1 if d.month>1 else 12, 1)
    meses.reverse()

    if request.args.get('export') == 'excel':
        return _export_economico_excel(filas, inicio, fin, total_periodo, total_ingresos)

    return render_template('reportes/economico.html',
        filas=filas, inicio=inicio, fin=fin, mes_str=mes_str,
        total_periodo=total_periodo, total_ingresos=total_ingresos,
        meses=meses, hoy=hoy)


def _export_economico_excel(filas, inicio, fin, total_compras, total_ingresos):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── HOJA 1: REGISTRO DE COMPRAS ──
    ws1 = wb.active
    ws1.title = 'REGISTRO DE COMPRAS'

    thin = Side(style='thin', color='CCCCCC')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor='1E293B')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_fmt   = '"S/."#,##0.00'

    # Título
    ws1.merge_cells('A1:K1')
    titulo = f'REPORTE ECONÓMICO — REGISTRO DE COMPRAS — {inicio.strftime("%B %Y").upper()}'
    ws1['A1'] = titulo
    ws1['A1'].font = Font(bold=True, size=12, color='1E293B')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1['A1'].fill = PatternFill('solid', fgColor='F1F5F9')
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells('A2:K2')
    ws1['A2'] = f'Período: {inicio.strftime("%d/%m/%Y")} al {fin.strftime("%d/%m/%Y")}'
    ws1['A2'].font = Font(size=9, color='64748B')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 16

    # Cabecera
    headers = ['ITEM','FECHA DE ADQ.','DETALLE','UND. MEDIDA','CTD.','P.U.','SUBTOTAL','TOTAL','RESPONSABLE','COMPROBANTE','RAZÓN SOCIAL']
    widths   = [6, 12, 38, 12, 8, 10, 12, 12, 20, 16, 24]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = bord
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.row_dimensions[3].height = 28

    # Datos
    fill_alt = PatternFill('solid', fgColor='F8FAFC')
    for r, f in enumerate(filas, 4):
        vals = [
            f['num'], f['fecha'].strftime('%d/%m/%Y') if f['fecha'] else '',
            f['detalle'], f['unidad'],
            f['cantidad'], f['precio_unit'],
            f['subtotal'], f['total'],
            f['responsable'], f['comprobante'], f['razon_social']
        ]
        row_fill = fill_alt if r % 2 == 0 else None
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.border = bord
            c.alignment = Alignment(
                horizontal='center' if col in (1,5) else 'right' if col in (6,7,8) else 'left',
                vertical='center')
            if col in (6,7,8) and isinstance(val, (int,float)) and val is not None:
                c.number_format = num_fmt
            if row_fill: c.fill = row_fill
        ws1.row_dimensions[r].height = 15

    # Fila total
    tr = len(filas) + 4
    ws1.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, size=10)
    ws1.cell(row=tr, column=7, value=f'=SUM(G4:G{tr-1})').font = Font(bold=True)
    ws1.cell(row=tr, column=7).number_format = num_fmt
    ws1.cell(row=tr, column=8, value=total_compras).font = Font(bold=True, color='DC2626')
    ws1.cell(row=tr, column=8).number_format = num_fmt
    for col in range(1, 12):
        ws1.cell(row=tr, column=col).fill = PatternFill('solid', fgColor='FEF2F2')
        ws1.cell(row=tr, column=col).border = bord
    ws1.freeze_panes = 'A4'

    # ── HOJA 2: RESUMEN INGRESOS/EGRESOS ──
    ws2 = wb.create_sheet('RESUMEN')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 16

    titulos_resumen = [
        ('RESUMEN DEL PERÍODO', None, '1E293B', 13),
        (f'{inicio.strftime("%B %Y").upper()}', None, '64748B', 10),
        ('', None, None, 12),
        ('INGRESOS (Ventas)', total_ingresos, '166534', 11),
        ('EGRESOS (Compras)', total_compras, 'DC2626', 11),
        ('BALANCE', total_ingresos - total_compras,
         '1D4ED8' if total_ingresos >= total_compras else 'DC2626', 11),
    ]
    for row, (label, val, color, size) in enumerate(titulos_resumen, 1):
        c1 = ws2.cell(row=row, column=1, value=label)
        c1.font = Font(bold=(val is not None), size=size,
                       color=color if color else '1E293B')
        c1.alignment = Alignment(vertical='center')
        ws2.row_dimensions[row].height = 22
        if val is not None:
            c2 = ws2.cell(row=row, column=2, value=val)
            c2.font = Font(bold=True, size=size, color=color)
            c2.number_format = num_fmt
            c2.alignment = Alignment(horizontal='right', vertical='center')

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    resp = make_response(output.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=reporte_economico_{inicio.strftime("%Y_%m")}.xlsx'
    return resp


# ──────────────────────────────────────
#  REPORTE DIARIO DE VENTAS (Kardex)
# ──────────────────────────────────────
@reportes_bp.route('/diario')
@login_required
def diario():
    from models import ProductoCarta, CategoriaCarta, ItemVenta, VentaDiaria, Producto, KardexComedor
    from sqlalchemy import func

    hoy     = date.today()
    fecha_str = request.args.get('fecha', hoy.strftime('%Y-%m-%d'))
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except:
        fecha = hoy

    # Todos los productos de carta activos
    productos = ProductoCarta.query.filter_by(activo=True).order_by(
        ProductoCarta.categoria_id, ProductoCarta.orden
    ).all()

    # Ventas del día — cantidades vendidas por producto
    ventas_dia = db.session.query(
        ItemVenta.producto_carta_id,
        func.sum(ItemVenta.cantidad).label('vendido')
    ).join(VentaDiaria).filter(
        VentaDiaria.fecha == fecha
    ).group_by(ItemVenta.producto_carta_id).all()

    vendido_map = {v.producto_carta_id: int(v.vendido or 0) for v in ventas_dia}

    # Construir filas del reporte
    filas = []
    for p in productos:
        vendido    = vendido_map.get(p.id, 0)
        # Stock actual del producto en almacén (si está vinculado)
        stock_actual = 0
        if p.producto_almacen_id and p.descuenta_inventario:
            prod_alm = Producto.query.get(p.producto_almacen_id)
            if prod_alm:
                stock_actual = prod_alm.stock_actual or 0
        # Stock inicial = stock actual + lo que se vendió hoy
        stock_inicial = stock_actual + vendido
        stock_final   = stock_actual

        filas.append({
            'id':            p.id,
            'nombre':        p.nombre,
            'categoria':     p.categoria_carta.nombre if p.categoria_carta else '—',
            'stock_inicial': stock_inicial,
            'vendido':       vendido,
            'stock_final':   stock_final,
            'precio':        p.precio,
            'total_venta':   round(vendido * p.precio, 2),
            'tiene_stock':   bool(p.producto_almacen_id and p.descuenta_inventario),
        })

    # Solo mostrar productos que tuvieron movimiento O que tienen stock vinculado
    filas_activas = [f for f in filas if f['vendido'] > 0 or f['tiene_stock']]

    # Totales
    total_ingresos  = sum(f['total_venta'] for f in filas)
    total_productos = sum(f['vendido'] for f in filas)

    # Resumen por empresa (para el lado derecho como en la foto)
    from models import EmpresaTuristica
    ventas_empresas = db.session.query(
        VentaDiaria.empresa_id,
        func.sum(VentaDiaria.total).label('total'),
        func.sum(VentaDiaria.num_pax).label('pax')
    ).filter(VentaDiaria.fecha == fecha).group_by(VentaDiaria.empresa_id).all()

    empresas_map = {e.id: e for e in EmpresaTuristica.query.all()}
    resumen_empresas = []
    for v in ventas_empresas:
        nombre = empresas_map[v.empresa_id].nombre if v.empresa_id and v.empresa_id in empresas_map else 'Privado'
        resumen_empresas.append({
            'nombre': nombre,
            'total': float(v.total or 0),
            'pax': int(v.pax or 0),
        })

    # Exportar
    if request.args.get('export') == 'excel':
        return _export_diario_excel(filas_activas, fecha, resumen_empresas, total_ingresos)

    return render_template('reportes/diario.html',
        filas=filas, filas_activas=filas_activas,
        fecha=fecha, fecha_str=fecha_str, hoy=hoy,
        total_ingresos=total_ingresos, total_productos=total_productos,
        resumen_empresas=resumen_empresas)


def _export_diario_excel(filas, fecha, resumen_empresas, total_ingresos):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import make_response

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte Diario'

    thin  = Side(style='thin', color='D1D5DB')
    bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
    blk   = PatternFill('solid', fgColor='1E293B')
    white = Font(bold=True, color='FFFFFF', size=10)
    ctr   = Alignment(horizontal='center', vertical='center')

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = f'REPORTE DIARIO DE VENTAS — REST. TCO. MARANGANI'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Fecha: {fecha.strftime("%d/%m/%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=10, color='64748B')
    ws.row_dimensions[2].height = 16

    # Cabecera
    hdrs  = ['DETALLE', 'STOCK INICIAL', 'SALIDA', 'STOCK FINAL', 'P.U. S/.', 'TOTAL S/.']
    widths= [28, 14, 12, 14, 10, 12]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = white; c.fill = blk; c.alignment = ctr; c.border = bord
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = 'A4'

    fill_alt = PatternFill('solid', fgColor='F8FAFC')
    fill_0   = PatternFill('solid', fgColor='FFFFFF')
    for r, f in enumerate(filas, 4):
        vals = [f['nombre'], f['stock_inicial'], f['vendido'],
                f['stock_final'], f['precio'], f['total_venta']]
        row_fill = fill_alt if r % 2 == 0 else fill_0
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = bord
            c.fill = row_fill
            c.alignment = Alignment(
                horizontal='center' if col in (2,3,4) else
                'right' if col in (5,6) else 'left',
                vertical='center')
            if col in (5,6):
                c.number_format = '"S/."#,##0.00'
            if col == 3 and val and val > 0:
                c.font = Font(bold=True, color='DC2626')
        ws.row_dimensions[r].height = 15

    # Fila total
    tr = len(filas) + 4
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=tr, column=3, value=f'=SUM(C4:C{tr-1})').font = Font(bold=True)
    ws.cell(row=tr, column=6, value=total_ingresos).font = Font(bold=True, color='166534')
    ws.cell(row=tr, column=6).number_format = '"S/."#,##0.00'
    for col in range(1,7):
        ws.cell(row=tr, column=col).fill = PatternFill('solid', fgColor='F0FDF4')
        ws.cell(row=tr, column=col).border = bord

    # Hoja 2: resumen por empresa
    ws2 = wb.create_sheet('Resumen empresas')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 10
    ws2.append(['EMPRESA', 'TOTAL S/.', 'PAX'])
    for cell in ws2[1]: cell.font = Font(bold=True); cell.fill = blk; cell.font = white
    for e in resumen_empresas:
        ws2.append([e['nombre'], e['total'], e['pax']])

    out = io.BytesIO(); wb.save(out); out.seek(0)
    resp = make_response(out.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=reporte_diario_{fecha.strftime("%Y%m%d")}.xlsx'
    return resp
